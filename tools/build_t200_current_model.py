"""Build the bundled T200 current model from BlueRobotics performance data.

This is a *one-time / occasional* developer tool. It converts the official
BlueRobotics T200 public performance spreadsheet into a compact JSON lookup
table (``control/t200_current_model.json``) that the runtime uses to predict
thruster current draw from commanded PWM and supply voltage.

The runtime never reads the xlsx (no openpyxl/pandas dependency on the Pi); it
only loads the generated JSON.

Usage:
    python -m tools.build_t200_current_model \
        --xlsx "C:/path/to/T200-Public-Performance-Data-10-20V-September-2019.xlsx"

The spreadsheet layout (one sheet per voltage, e.g. "12 V") has columns:
    PWM (us) | RPM | Current (A) | Voltage (V) | Power (W) | Force (Kg f) | Eff
We keep only PWM and Current; current is a positive magnitude (draw) in both
forward (PWM > 1500) and reverse (PWM < 1500) directions, and ~0 at neutral.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from typing import Dict, List, Tuple


def _sheet_voltage(name: str) -> float | None:
    m = re.search(r"(\d+(?:\.\d+)?)\s*V", str(name), re.IGNORECASE)
    return float(m.group(1)) if m else None


def extract(xlsx_path: str) -> Dict[float, List[Tuple[int, float]]]:
    import openpyxl  # local import: dev-only dependency

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: Dict[float, List[Tuple[int, float]]] = {}
    for name in wb.sheetnames:
        voltage = _sheet_voltage(name)
        if voltage is None:
            continue  # skip "READ ME FIRST" etc.
        ws = wb[name]
        pts: List[Tuple[int, float]] = []
        for r in range(2, ws.max_row + 1):
            pwm = ws.cell(row=r, column=1).value
            cur = ws.cell(row=r, column=3).value
            if pwm is None or cur is None:
                continue
            pts.append((int(round(float(pwm))), round(abs(float(cur)), 3)))
        pts.sort(key=lambda p: p[0])
        if pts:
            out[voltage] = pts
    return out


def main() -> None:
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_out = os.path.join(here, "control", "t200_current_model.json")

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, help="path to the BlueRobotics T200 xlsx")
    ap.add_argument("--out", default=default_out, help=f"output JSON (default: {default_out})")
    args = ap.parse_args()

    curves = extract(args.xlsx)
    if not curves:
        raise SystemExit("no voltage sheets found in workbook")

    voltages = sorted(curves)
    payload = {
        "meta": {
            "source": os.path.basename(args.xlsx),
            "description": "T200 current draw (A) vs PWM (us) per supply voltage (V).",
            "pwm_neutral_us": 1500,
            "note": "current is a positive magnitude in both directions; ~0 at neutral.",
        },
        "voltages": voltages,
        "curves": [
            {
                "voltage": v,
                "pwm_us": [p for p, _ in curves[v]],
                "current_a": [c for _, c in curves[v]],
            }
            for v in voltages
        ],
    }

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(payload, f, separators=(",", ":"))
    sizes = ", ".join(f"{v:g}V:{len(curves[v])}pts" for v in voltages)
    print(f"wrote {args.out}  ({sizes})")


if __name__ == "__main__":
    main()
